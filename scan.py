# -*- coding: UTF-8 -*-
'''
Desprition:
    自动化批量nmap使用

Author:
    Sp4rkW   https://sp4rkw.blog.csdn.net/

Modify:2019-08-05 14:42:20
'''

# -*- coding: utf-8 -*-
import datetime,time,openpyxl,sys,getopt,nmap,threading
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sem=threading.Semaphore(4) #限定最大线程数

'''
Desprition:
    函数 ReadExcel 用于将 nmap 数据写入 xlsl 文件，并且做自适应宽度调整

Parameters:
    param nmap扫描之后的结果
    url 目标站点url
    starttime 开始的时间

Returns:
    Null

Modify:2019-08-05 15:05:01
'''
def WriteExcel(param,url,starttime):
    wb = Workbook()
    sheet = wb.active
    sheet.title = url
    len1 = len(param)
    for i in range(0, len1):
        demo = param[i].split(';')
        len2 = len(demo)
        for j in range(0, len2):
            sheet.cell(row=i+1, column=j+1, value=str(demo[j]))
    col_width = []   #记录每一列的宽度
    i = 0
    for col in sheet.columns:
        for j in range(len(col)):
            if j == 0:
	         	# 数组增加一个元素
	            col_width.append(len(str(col[j].value)))
            else:
                # 获得每列中的内容的最大宽度
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1
    #设置列宽
    for j in range(len(col_width)):
        # 根据列的数字返回字母
        col_letter = get_column_letter(j+1)
        # 当宽度大于100，宽度设置为100
        if col_width[j] > 100:
            sheet.column_dimensions[col_letter].width = 100
        # 只有当宽度大于10，才设置列宽
        elif col_width[j] > 7:
            sheet.column_dimensions[col_letter].width = col_width[j] + 2
    wb.save('./'+url+'.xlsx')
    print('#*# 文件：'+url+'.xlsx 已经输出完毕')
    endtime = datetime.datetime.now()
    print ('耗时：'+str((endtime - starttime).seconds)+'s')


'''
Desprition:
    函数 NmapExcel 调用 python-nmap 使用 nmap 扫描

Parameters:
    Null

Returns:
    Null

Modify:2019-08-05 15:05:22
'''

def NmapExcel(param1,param2,url):
    with sem:
        nm = nmap.PortScanner() #创建nmap接口
        # try:
        starttime = datetime.datetime.now()
        nm.scan(hosts=url, ports=param1, arguments=' -Pn -sS -sU '+param2)
        nmap_data = nm.csv().split('\n')
        WriteExcel(nmap_data,url,starttime)
        # except Exception:
        #     print('扫描出错，当前扫描进行到：'+url)




'''
Desprition:
    函数 StartFunc 用于输出相关参数的意义

Parameters:
    Null

Returns:
    Null

Modify:2019-08-05 15:08:52
'''
def StartFunc(argv): 
    parameter1 = ''
    parameter2 = ''
    try:
        opts, args = getopt.getopt(sys.argv[1:],"hp:o:",["port=","other="])
    except getopt.GetoptError:
        print('python ./scan.py -p <需要扫描端口> -o <其他nmap参数>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('python ./scan.py -p <需要扫描端口> -o <其他nmap参数>')
            sys.exit()
        elif opt in ("-p", "--port"):
            parameter1 = arg
        elif opt in ("-o", "--other"):
            parameter2 = arg
    print('''
         )\____________________________|¯¯¯¯¯|__,,,,,. 
      )_         ||||||||||||||||||||||||||||||||||||||||||||||||||||||||            |    | 
        /                    _________________         __|__| 
       /}}}}}}}}}}/(__\¯\¯                              |____| 
      /}}}}}}}}}}}/}}___\| 
     /}}}}}}}}}}}}}}/ 
    /}}}}}}}}}}}}}}/ 
   /}}}}}}}}}}}}}}/ 
  /}}}}}}}}}}}}}}/ 
 /}}}}}}}}}}}}}}/ 
/}}}}}}}}}}}}}}/ 
¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯

NmapToExcel version 1.1.0
time 2019-8-5
    ''')
    print('#*# NmapToExcel 正在运行')
    data = open('url.txt','r') #url.txt存储的是目标站点ip/url
    website_urls = data.readlines()
    if website_urls:
        for url in website_urls:
            threading.Thread(target=NmapExcel, args=(parameter1,parameter2,url.replace('\n',''))).start()
    else:
            print("未读取到目标站点，请先添加目标")
    data.close()

    



if __name__ == "__main__":
    StartFunc(sys.argv[1:])




